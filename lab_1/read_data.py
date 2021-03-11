from openpyxl import load_workbook


def read_number_train():
    X = []
    table = load_workbook('../number_train.xlsx')
    table_sheet_names = table.sheetnames
    for sheet in range(len(table.sheetnames)):
        X.append([])
        for i in range(5):
            for j in range(7):
                X[len(X) - 1].append(table.get_sheet_by_name(str(table_sheet_names[sheet])).cell(row=j+1, column=i+1).value)
    return X


def read_number_data():
    X = []
    table = load_workbook('../number_data.xlsx')
    table_sheet = table.get_sheet_by_name("0")
    for i in range(5):
        for j in range(7):
            X.append(table_sheet.cell(row=j+1, column=i+1).value)
    return X
