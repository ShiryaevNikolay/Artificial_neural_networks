import numpy as np
from openpyxl import *


# Активационная функция (ступенчатая функция)
def activation_fun(net_y, h):
    # if net_y > h:
    #     return 1
    # elif net_y <= h:
    #     return 0
    return 1 / (1 + np.exp(-net_y))


# функция NET
def net(x, w):
    net_y = 0
    for i in range(len(x)):
        net_y += x[i] * w[i]
    return net_y


def init_weight(length):
    weight = []
    for i in range(length):
        weight.append(np.random.randint(-1, 1))
    return weight


def weight_sheet_exist(id_perceptron, weight_file_path="../number_weight"):
    exist = False
    weight_file = load_workbook(weight_file_path)
    if str(id_perceptron) in weight_file.sheetnames:
        exist = True
    weight_file.close()
    return exist

def cell_empty(id_perceptron, h_file_path="../number_h"):
    is_empty = True
    h_file = load_workbook(h_file_path)
    sheet = h_file.get_sheet_by_name("h")
    if sheet.cell(row=id_perceptron + 1, column=1).value:
        is_empty = False
    h_file.save(h_file_path)
    h_file.close()
    return is_empty
