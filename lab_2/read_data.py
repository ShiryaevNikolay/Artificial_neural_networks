from openpyxl import *

table_file = '../number_weight.xlsx'
table_target_file = '../number_target.xlsx'


def read_number_train():
    X = []
    table = load_workbook('../number_train.xlsx')
    table_sheet_names = table.sheetnames
    for sheet in range(len(table.sheetnames)):
        X.append([])
        for i in range(5):
            for j in range(7):
                X[len(X) - 1].append(table.get_sheet_by_name(str(table_sheet_names[sheet])).cell(row=j+1, column=i+1).value)
    return X


def read_number_data():
    X = []
    table = load_workbook('../number_data.xlsx')
    table_sheet = table.get_sheet_by_name("0")
    for i in range(5):
        for j in range(7):
            X.append(table_sheet.cell(row=j+1, column=i+1).value)
    return X


def read_target(number):
    table = load_workbook(table_target_file)
    target = []
    for i in range(10):
        target.append(table.get_sheet_by_name(str(1)).cell(row=i+1, column=number+1).value)
    return target


def read_weight(id_perceptron, count, weight_file_path="../number_weight"):
    weights = []
    weight_file = load_workbook(weight_file_path)
    sheet = weight_file.get_sheet_by_name(str(id_perceptron))
    for i in range(count):
        weights.append(sheet.cell(row=i + 1, column=1).value)
    weight_file.close()
    return weights


def read_h(id_perceptron, h_file_path="../number_h"):
    h_file = load_workbook(h_file_path)
    sheet = h_file.get_sheet_by_name("h")
    h = sheet.cell(row=id_perceptron + 1, column=1).value
    h_file.close()
    return h
