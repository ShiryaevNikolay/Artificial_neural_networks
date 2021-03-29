import numpy as np
from openpyxl import *


def create_weight_file(weight_file_path="../number_weight"):
    weight_file = Workbook()
    weight_file.save(weight_file_path)
    weight_file.close()


def create_h_file(h_file_path="../number_h"):
    h_file = Workbook()
    h_file.create_sheet("h")
    h_file.save(h_file_path)
    h_file.close()


def init_weight_and_save_to_file(id_perceptron, count, weight_file_path="../number_weight"):
    weight_file = load_workbook(weight_file_path)
    weight_file.create_sheet(str(id_perceptron))
    sheet = weight_file.get_sheet_by_name(str(id_perceptron))
    for i in range(count):
        sheet.cell(row=i + 1, column=1).value = np.random.randint(-1, 1)
    weight_file.save(weight_file_path)
    weight_file.close()


def init_h(id_perceptron, h, h_file_path="../number_h"):
    h_file = load_workbook(h_file_path)
    sheet = h_file.get_sheet_by_name("h")
    sheet.cell(row=id_perceptron + 1, column=1).value = h
    h_file.save(h_file_path)
    h_file.close()


def write_weight(id_perceptron, weights, weight_file_path="../number_weight"):
    weight_file = load_workbook(weight_file_path)
    sheet = weight_file.get_sheet_by_name(str(id_perceptron))
    for i in range(len(weights)):
        sheet.cell(row=i + 1, column=1).value = weights[i]
    weight_file.save(weight_file_path)
    weight_file.close()


def write_h(id_perceptron, h, h_file_path="../number_h"):
    h_file = load_workbook(h_file_path)
    sheet = h_file.get_sheet_by_name("h")
    sheet.cell(row=id_perceptron + 1, column=1).value = h
    h_file.save(h_file_path)
    h_file.close()
