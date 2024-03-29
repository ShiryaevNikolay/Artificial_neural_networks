import numpy as np
from openpyxl import *


def create_weight_file(weight_file_path="../mnist_weight.xlsx"):
    weight_file = Workbook()
    weight_file.save(weight_file_path)
    weight_file.close()


def init_weight_and_save_to_file(id_perceptron, count, weight_file_path="../mnist_weight.xlsx"):
    weight_file = load_workbook(weight_file_path)
    weight_file.create_sheet(str(id_perceptron))
    sheet = weight_file.get_sheet_by_name(str(id_perceptron))
    for i in range(28):
        for j in range(28):
            sheet.cell(row=i + 1, column=j + 1).value = np.random.uniform(0, 1)
    weight_file.save(weight_file_path)
    weight_file.close()


def write_weight(id_perceptron, weights, weight_file_path="../mnist_weight.xlsx"):
    weight_file = load_workbook(weight_file_path)
    sheet = weight_file.get_sheet_by_name(str(id_perceptron))
    for i in range(28):
        for j in range(28):
            sheet.cell(row=i + 1, column=j + 1).value = weights[i][j]
    weight_file.save(weight_file_path)
    weight_file.close()
