from openpyxl import *

table_file = '../mnist_weight.xlsx'


def read_weight(id_perceptron, count, weight_file_path="../mnist_weight.xlsx"):
    weight_file = load_workbook(weight_file_path)
    weights = []
    sheet = weight_file.get_sheet_by_name(str(id_perceptron))
    for i in range(28):
        weights.append([])
        for j in range(28):
            weights[i].append(sheet.cell(row=i + 1, column=j + 1).value)
    weight_file.close()
    return weights


def get_sheets_weight(weight_file_path="../mnist_weight.xlsx"):
    weight_file = load_workbook(weight_file_path)
    sheets = weight_file.sheetnames
    weight_file.close()
    return sheets
